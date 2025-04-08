import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


def make_safe(name):
    return re.sub(r'\W+', '_', str(name)).strip('_')

def clean_data(df):
    df.ffill(inplace=True)
    df.dropna(subset=['Registration Id', 'Student', 'Present %', 'Course [Course Code]'], inplace=True)
    df['Present %'] = pd.to_numeric(df['Present %'], errors='coerce')
    df['Overall Present %'] = pd.to_numeric(df['Overall Present %'], errors='coerce')
    df.dropna(subset=['Present %', 'Overall Present %'], inplace=True)
    return df


def extract_subject_codes(filepath):
    df = pd.read_excel(filepath)
    df = clean_data(df)
    df['Subject Code'] = df['Course [Course Code]'].apply(
        lambda x: re.search(r'\[(.*?)\]', str(x)).group(1) if pd.notna(x) and '[' in str(x) else 'Unknown'
    )
    df['Subject Name'] = df['Course [Course Code]'].apply(
        lambda x: str(x).split(' [')[0] if pd.notna(x) else 'Unknown'
    )
    unique = df[['Subject Code', 'Subject Name']].drop_duplicates()
    return list(unique.itertuples(index=False, name=None)), df

def get_university_header(subject_code):
    styles = getSampleStyleSheet()
    return Paragraph(
        f"""
        <b>NSHM Knowledge Campus, Durgapur</b><br/>
        <b>Subject Eligibility Report</b><br/><b>Subject Code:</b> {subject_code}
        """,
        ParagraphStyle(
            name="UniversityHeader",
            parent=styles["Title"],
            fontSize=14,
            leading=18,
            alignment=1,
            spaceAfter=20,
        )
    )


def create_section_table(section_df, styles):
    normal_style = styles["Normal"]
    section_df = section_df[[
        'Student', 'Registration Id', 'Present %', 'Overall Present %'
    ]]

    total_width = A4[0] - 80
    col_width = total_width / len(section_df.columns)
    col_widths = [col_width] * len(section_df.columns)

    data = [[Paragraph(str(col), normal_style) for col in section_df.columns]]
    for row in section_df.itertuples(index=False):
        data.append([Paragraph(str(cell), normal_style) for cell in row])

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E5E5E5")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")])
    ]))
    return table


def export_pdf(subject_code, df, folder_path):
    filename = os.path.join(folder_path, f"{subject_code}_eligibility.pdf")
    pdf = SimpleDocTemplate(filename, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    grouped = df.groupby(['Programme', 'Programme Section'])
    for i, ((programme, section), section_df) in enumerate(grouped):
        if i > 0:
            elements.append(PageBreak())

        section_title = Paragraph(
            f"<b>Programme:</b> {programme} | <b>Section:</b> {section}",
            ParagraphStyle(name="SectionHeader", fontSize=12, leading=14, spaceBefore=12, spaceAfter=8)
        )
        elements.append(get_university_header(subject_code))
        elements.append(Spacer(1, 12))
        elements.append(section_title)
        table = create_section_table(section_df, styles)
        elements.extend([table, Spacer(1, 16)])

    pdf.build(elements)


def export_combined_pdf(df, selected_subject_codes, output_folder_path):
    filename = os.path.join(output_folder_path, "Combined_Subject_Report.pdf")
    doc = SimpleDocTemplate(filename, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    for idx, code in enumerate(selected_subject_codes):
        sub_df = df[(df['Subject Code'] == code) & (df['Subject Eligible'])].copy()
        if sub_df.empty:
            continue

        if idx > 0:
            elements.append(PageBreak())

        
        subject_name = sub_df['Subject Name'].iloc[0]
        grouped = sub_df.groupby(['Programme', 'Programme Section'])

        for i, ((programme, section), section_df) in enumerate(grouped):
            elements.append(PageBreak() if i > 0 else Spacer(1, 12))
            section_title = Paragraph(
                f"<b>Programme:</b> {programme} <br/><b>Section:</b> {section}",
                ParagraphStyle(name="SectionHeader", fontSize=12, leading=14, spaceBefore=12, spaceAfter=8)
            )
            elements.append(get_university_header(code))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"<b>Subject:</b> {subject_name} ({code})", styles['Heading2']))
            elements.append(Spacer(1, 12))
            elements.append(section_title)
            table = create_section_table(section_df, styles)
            elements.extend([table, Spacer(1, 16)])

    doc.build(elements)


def process_file(df, selected_subject_codes, output_folder_path, combine_subjects=False, overall_threshold=75, subjectwise_threshold=75):
    df['Present %'] = pd.to_numeric(df['Present %'], errors='coerce')
    df['Overall Present %'] = pd.to_numeric(df['Overall Present %'], errors='coerce')

    overall_eligible = df[['Registration Id', 'Overall Present %']].drop_duplicates()
    overall_eligible['Eligible for All Subjects'] = overall_eligible['Overall Present %'] >= overall_threshold
    df = df.merge(overall_eligible[['Registration Id', 'Eligible for All Subjects']], on='Registration Id', how='left')
    df['Subject Eligible'] = df.apply(
        lambda row: True if row['Eligible for All Subjects'] else row['Present %'] >= subjectwise_threshold, axis=1
    )

    df['Subject Code'] = df['Course [Course Code]'].apply(
        lambda x: re.search(r'\[(.*?)\]', str(x)).group(1) if pd.notna(x) and '[' in str(x) else 'Unknown'
    )
    df['Subject Name'] = df['Course [Course Code]'].apply(
        lambda x: str(x).split(' [')[0] if pd.notna(x) else 'Unknown'
    )
    df['Subject Code Safe'] = df['Subject Code'].apply(make_safe)

    if combine_subjects:
        export_combined_pdf(df, selected_subject_codes, output_folder_path)

    summary = df.groupby(['Subject Code', 'Subject Name']).agg(
        Total_Students=('Registration Id', 'nunique'),
        Eligible_Students=('Subject Eligible', lambda x: df.loc[x.index, 'Registration Id'][x].nunique())
    ).reset_index()
    summary['Eligibility %'] = round((summary['Eligible_Students'] / summary['Total_Students']) * 100, 2)
    summary = summary[summary['Subject Code'].isin(selected_subject_codes)]

    output_file = os.path.join(output_folder_path, "subjectwise_eligibility.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for subject_code in df['Subject Code'].unique():
            if subject_code not in selected_subject_codes:
                continue
            sub_df = df[(df['Subject Code'] == subject_code) & (df['Subject Eligible'])].copy()
            if sub_df.empty:
                continue
            export_df = sub_df[[
                'Student', 'Registration Id', 'Course [Course Code]', 'Present %', 'Overall Present %', 'Programme', 'Programme Section'
            ]]
            export_pdf(subject_code, export_df.sort_values(by='Programme Section'), output_folder_path)
            export_df.to_excel(writer, sheet_name=subject_code[:31], index=False)
        summary.to_excel(writer, sheet_name="Dashboard", index=False)

    wb = load_workbook(output_file)
    ws = wb["Dashboard"]
    chart = BarChart()
    chart.title = "Eligible Students per Subject"
    chart.x_axis.title = "Subject Code"
    chart.y_axis.title = "Eligible Students"
    data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 10
    ws.add_chart(chart, "G2")

    rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                          mid_type='num', mid_value=60, mid_color='FFEB84',
                          end_type='num', end_value=75, end_color='63BE7B')
    ws.conditional_formatting.add(f'E2:E{ws.max_row}', rule)

    wb.save(output_file)
    return output_file