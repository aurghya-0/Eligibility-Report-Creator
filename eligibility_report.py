import pandas as pd
import re
import os
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import threading

selected_subjects = set()
subject_code_name_map = {}
input_filepath = ""
output_folder_path = ""

def make_safe(name):
    return re.sub(r'\W+', '_', str(name)).strip('_')

def export_pdf(subject_code, df, folder_path):
    filename = os.path.join(folder_path, f"{subject_code}_eligibility.pdf")
    pdf = SimpleDocTemplate(filename, pagesize=landscape(A4))
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph(f"Eligibility List for Subject: <b>{subject_code}</b>", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))

    data = [df.columns.tolist()] + df.values.tolist()
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))
    elements.append(table)
    pdf.build(elements)

def clean_data(df):
    df.ffill(inplace=True)
    df.dropna(subset=['Registration Id', 'Student', 'Present %', 'Course [Course Code]'], inplace=True)
    df['Present %'] = pd.to_numeric(df['Present %'], errors='coerce')
    df['Overall Present %'] = pd.to_numeric(df['Overall Present %'], errors='coerce')
    df.dropna(subset=['Present %', 'Overall Present %'], inplace=True)
    return df

def extract_subject_codes(filepath):
    df = pd.read_excel(filepath)
    df = clean_data(df)
    df['Subject Code'] = df['Course [Course Code]'].apply(
        lambda x: re.search(r'\[(.*?)\]', str(x)).group(1) if pd.notna(x) and '[' in str(x) else 'Unknown'
    )
    df['Subject Name'] = df['Course [Course Code]'].apply(
        lambda x: str(x).split(' [')[0] if pd.notna(x) else 'Unknown'
    )
    unique = df[['Subject Code', 'Subject Name']].drop_duplicates()
    return list(unique.itertuples(index=False, name=None)), df

def process_file(df, selected_subject_codes):
    df['Present %'] = pd.to_numeric(df['Present %'], errors='coerce')
    df['Overall Present %'] = pd.to_numeric(df['Overall Present %'], errors='coerce')

    overall_eligible = df[['Registration Id', 'Overall Present %']].drop_duplicates()
    overall_eligible['Eligible for All Subjects'] = overall_eligible['Overall Present %'] >= 75
    df = df.merge(overall_eligible[['Registration Id', 'Eligible for All Subjects']], on='Registration Id', how='left')
    df['Subject Eligible'] = df.apply(
        lambda row: True if row['Eligible for All Subjects'] else row['Present %'] >= 75, axis=1
    )

    df['Subject Code'] = df['Course [Course Code]'].apply(
        lambda x: re.search(r'\[(.*?)\]', str(x)).group(1) if pd.notna(x) and '[' in str(x) else 'Unknown'
    )
    df['Subject Name'] = df['Course [Course Code]'].apply(
        lambda x: str(x).split(' [')[0] if pd.notna(x) else 'Unknown'
    )

    df['Subject Code Safe'] = df['Subject Code'].apply(make_safe)

    summary = df.groupby(['Subject Code', 'Subject Name']).agg(
        Total_Students=('Registration Id', 'nunique'),
        Eligible_Students=('Subject Eligible', lambda x: df.loc[x.index, 'Registration Id'][x].nunique())
    ).reset_index()
    summary['Eligibility %'] = round((summary['Eligible_Students'] / summary['Total_Students']) * 100, 2)

    output_file = os.path.join(output_folder_path, "subjectwise_eligibility.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for subject_code in df['Subject Code'].unique():
            if subject_code not in selected_subject_codes:
                continue
            sub_df = df[(df['Subject Code'] == subject_code) & (df['Subject Eligible'])].copy()
            if sub_df.empty:
                continue
            export_df = sub_df[[ 'Student', 'Registration Id', 'Course [Course Code]', 'Present %', 'Overall Present %', 'Programme Section']]
            folder_path = output_folder_path
            export_pdf(subject_code, export_df.sort_values(by='Programme Section'), folder_path)

            sheet_name = subject_code[:31] if subject_code else "Unknown"
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
        summary.to_excel(writer, sheet_name="Dashboard", index=False)

    wb = load_workbook(output_file)
    ws = wb["Dashboard"]
    chart = BarChart()
    chart.title = "Eligible Students per Subject"
    chart.x_axis.title = "Subject Code"
    chart.y_axis.title = "Eligible Students"
    data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 10
    ws.add_chart(chart, "G2")

    rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                          mid_type='num', mid_value=60, mid_color='FFEB84',
                          end_type='num', end_value=75, end_color='63BE7B')
    ws.conditional_formatting.add(f'E2:E{ws.max_row}', rule)

    wb.save(output_file)
    return output_file

def choose_excel():
    global input_filepath, subject_code_name_map
    input_filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if input_filepath:
        subject_code_name_map.clear()
        subject_list, df = extract_subject_codes(input_filepath)
        subject_code_name_map.update({f"{code} - {name}": code for code, name in subject_list})
        populate_checkboxes(list(subject_code_name_map.keys()))
        status_label.config(text=f"Loaded {len(subject_list)} subjects")

def choose_output_folder():
    global output_folder_path
    output_folder_path = filedialog.askdirectory()
    if output_folder_path:
        status_label.config(text=f"Output Folder Selected: {output_folder_path}")

def populate_checkboxes(subject_list):
    for widget in checkbox_frame.winfo_children():
        widget.destroy()
    for sub in subject_list:
        var = BooleanVar()
        cb = Checkbutton(checkbox_frame, text=sub, variable=var, bg="#f5f5f5", anchor="w")
        cb.var = var
        cb.pack(fill="x", padx=5, pady=2)
        checkbox_vars[sub] = cb

def filter_checkboxes(*args):
    query = search_var.get().lower()
    for text, cb in checkbox_vars.items():
        if query in text.lower():
            cb.pack(fill="x", padx=5, pady=2)
        else:
            cb.pack_forget()

def generate_reports():
    if not input_filepath or not output_folder_path:
        messagebox.showwarning("Missing Info", "Please select input file and output folder.")
        return

    selected = [subject_code_name_map[text] for text, cb in checkbox_vars.items() if cb.var.get()]
    if not selected:
        messagebox.showwarning("No Selection", "Please select at least one subject.")
        return

    progress.start()
    threading.Thread(target=run_report_generation, args=(selected,), daemon=True).start()

def run_report_generation(selected):
    try:
        df = pd.read_excel(input_filepath)
        df = clean_data(df)
        output = process_file(df, selected)
        messagebox.showinfo("Success", f"Report saved to:\n{output}")
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        progress.stop()

root = Tk()
root.title("Eligibility Report Generator")
root.geometry("750x600")
root.configure(bg="#f5f5f5")

style = ttk.Style(root)
style.configure("TButton", font=("Segoe UI", 10), padding=6)
style.configure("TLabel", font=("Segoe UI", 10), background="#f5f5f5")

top_frame = Frame(root, bg="#f5f5f5")
top_frame.pack(pady=10)

ttk.Button(top_frame, text="Select Excel File", command=choose_excel).grid(row=0, column=0, padx=10)
ttk.Button(top_frame, text="Select Output Folder", command=choose_output_folder).grid(row=0, column=1, padx=10)
ttk.Button(top_frame, text="Generate Report", command=generate_reports).grid(row=0, column=2, padx=10)

status_label = Label(root, text="No file selected", bg="#f5f5f5", fg="gray")
status_label.pack()

search_var = StringVar()
search_var.trace("w", filter_checkboxes)
search_entry = Entry(root, textvariable=search_var, font=("Segoe UI", 10), width=50)
search_entry.pack(pady=5)
search_entry.insert(0, "Search subjects...")

checkbox_frame_container = Frame(root, bg="#f5f5f5")
checkbox_frame_container.pack(fill=BOTH, expand=True, padx=10, pady=10)

canvas = Canvas(checkbox_frame_container, bg="#f5f5f5")
scrollbar = Scrollbar(checkbox_frame_container, orient=VERTICAL, command=canvas.yview)
checkbox_frame = Frame(canvas, bg="#f5f5f5")

checkbox_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
canvas.create_window((0, 0), window=checkbox_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

def _on_mousewheel(event):
    canvas.yview_scroll(int(-1*(event.delta/120)), "units")

canvas.bind_all("<MouseWheel>", _on_mousewheel)

canvas.pack(side=LEFT, fill=BOTH, expand=True)
scrollbar.pack(side=RIGHT, fill=Y)

progress = ttk.Progressbar(root, mode='indeterminate')
progress.pack(fill=X, padx=10, pady=10)

checkbox_vars = {}

root.mainloop()
